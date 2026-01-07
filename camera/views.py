import cv2
import threading
import numpy as np
import json
import csv
import requests
from datetime import datetime, timedelta
from collections import defaultdict
from django.shortcuts import render
from django.http import StreamingHttpResponse, JsonResponse, HttpResponse
from django.views.decorators.http import condition
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import logging
import platform
from detector.yolo_detector import YOLODetector
from detector.traffic_controller import TrafficController
from detector.pedestrian_detector import PedestrianGestureDetector
from hardware.led_strip import LEDStripController
from camera.droidcam import DroidCamHandler

logger = logging.getLogger(__name__)

class VideoCamera:
    """
    Handles video capture from Raspberry Pi Camera Module 3
    with YOLO car detection
    Optimized for low-resource Raspberry Pi 5 environment
    """
    def __init__(self):
        self.video = None
        self.frame = None
        self.lock = threading.Lock()
        self.is_running = False
        self.thread = None
        self.picamera2 = None
        self.is_rpi = platform.machine() in ('armv7l', 'armv6l', 'aarch64')
        
        # Camera rotation (0, 90, 180, 270)
        self.rotation = 180  # Camera mounted upside down
        
        # YOLO detector
        self.detector = YOLODetector(model_name='yolov8n.pt')  # nano model for RPi
        self.detector_enabled = False
        self.car_count = 0
        
    def init_camera(self):
        """Initialize camera with native picamera2 for Raspberry Pi or OpenCV fallback"""
        try:
            # Try picamera2 first (works on Raspberry Pi)
            try:
                import sys
                logger.info(f"Python path: {sys.executable}")
                logger.info(f"Python version: {sys.version}")
                
                from picamera2 import Picamera2
                logger.info("picamera2 import successful")
                
                self.picamera2 = Picamera2()
                logger.info("Picamera2 object created")
                
                # Configure camera for balanced quality/performance
                config = self.picamera2.create_preview_configuration(
                    main={"format": "RGB888", "size": (640, 480)}
                )
                logger.info("Camera config created")
                
                self.picamera2.configure(config)
                logger.info("Camera configured")
                
                self.picamera2.start()
                logger.info("Camera started")
                
                self.is_running = True
                self.thread = threading.Thread(target=self._read_frames_picamera2)
                self.thread.daemon = True
                self.thread.start()
                logger.info("Picamera2 initialized successfully")
                return True
                
            except ImportError as e:
                logger.warning(f"picamera2 ImportError: {e}, trying OpenCV")
                import traceback
                logger.warning(traceback.format_exc())
            except Exception as e:
                logger.warning(f"picamera2 initialization failed: {e}, trying OpenCV")
                import traceback
                logger.warning(traceback.format_exc())
            
            # Fallback to OpenCV
            logger.info("Attempting OpenCV fallback")
            
            # Try multiple camera indices (0, 1, 2)
            camera_opened = False
            for cam_idx in [0, 1, 2]:
                logger.info(f"Trying camera index {cam_idx}")
                self.video = cv2.VideoCapture(cam_idx)
                
                if self.video.isOpened():
                    # Test if we can actually read a frame
                    ret, test_frame = self.video.read()
                    if ret and test_frame is not None:
                        logger.info(f"Camera {cam_idx} works - got test frame")
                        camera_opened = True
                        break
                    else:
                        logger.warning(f"Camera {cam_idx} opened but cannot read frames")
                        self.video.release()
                else:
                    logger.warning(f"Camera {cam_idx} failed to open")
            
            if not camera_opened:
                logger.error("Cannot open any camera with OpenCV (tried indices 0, 1, 2)")
                return False
            
            # Configure camera
            self.video.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            self.video.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            self.video.set(cv2.CAP_PROP_FPS, 30)
            
            # Read initial frame to warm up camera
            ret, self.frame = self.video.read()
            if ret:
                logger.info(f"Initial frame: {self.frame.shape}")
            
            self.is_running = True
            self.thread = threading.Thread(target=self._read_frames)
            self.thread.daemon = True
            self.thread.start()
            
            logger.info("OpenCV camera initialized (fallback)")
            return True
            
        except Exception as e:
            logger.error(f"Error initializing camera: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def _read_frames_picamera2(self):
        """Continuously read frames from picamera2"""
        try:
            while self.is_running:
                try:
                    request = self.picamera2.capture_request()
                    
                    # Get RGB data from the main stream
                    frame = request.make_array("main")
                    
                    # Convert RGB to BGR for OpenCV compatibility
                    frame_bgr = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
                    
                    # Apply rotation if needed (180 degrees for upside-down mount)
                    if self.rotation == 180:
                        frame_bgr = cv2.rotate(frame_bgr, cv2.ROTATE_180)
                    elif self.rotation == 90:
                        frame_bgr = cv2.rotate(frame_bgr, cv2.ROTATE_90_CLOCKWISE)
                    elif self.rotation == 270:
                        frame_bgr = cv2.rotate(frame_bgr, cv2.ROTATE_90_COUNTERCLOCKWISE)
                    
                    request.release()
                    
                    with self.lock:
                        self.frame = frame_bgr
                        
                except Exception as e:
                    logger.warning(f"Error capturing frame from picamera2: {e}")
                    import traceback
                    logger.warning(traceback.format_exc())
                    
        except Exception as e:
            logger.error(f"Picamera2 thread error: {e}")
            import traceback
            logger.error(traceback.format_exc())
        finally:
            if self.picamera2:
                try:
                    self.picamera2.stop()
                except:
                    pass
    
    def _read_frames(self):
        """Continuously read frames in background thread"""
        consecutive_failures = 0
        max_consecutive_failures = 30  # Allow 30 consecutive failures before giving up
        
        while self.is_running:
            ret, frame = self.video.read()
            
            if ret and frame is not None:
                # Apply rotation if needed (180 degrees for upside-down mount)
                if self.rotation == 180:
                    frame = cv2.rotate(frame, cv2.ROTATE_180)
                elif self.rotation == 90:
                    frame = cv2.rotate(frame, cv2.ROTATE_90_CLOCKWISE)
                elif self.rotation == 270:
                    frame = cv2.rotate(frame, cv2.ROTATE_90_COUNTERCLOCKWISE)
                    
                with self.lock:
                    self.frame = frame
                consecutive_failures = 0  # Reset counter on success
            else:
                consecutive_failures += 1
                if consecutive_failures == 1:  # Log only first failure
                    logger.warning(f"Failed to read frame from camera")
                
                if consecutive_failures >= max_consecutive_failures:
                    logger.error(f"Camera failed {max_consecutive_failures} times consecutively, stopping")
                    break
                    
                # Wait a bit before retrying
                import time
                time.sleep(0.1)
    
    def get_frame(self):
        """Get the latest frame"""
        with self.lock:
            if self.frame is not None:
                return self.frame.copy()
        return None
    
    def release(self):
        """Release camera resources"""
        self.is_running = False
        if self.thread:
            self.thread.join(timeout=5)
        if self.video:
            self.video.release()
        if self.picamera2:
            try:
                self.picamera2.stop()
            except:
                pass
        logger.info("Camera released")


# Global camera instance
camera = VideoCamera()
try:
    led_strip = LEDStripController(num_pixels=8, brightness=64)
except Exception as exc:
    led_strip = None
    logger.warning(f"LED strip not initialized: {exc}")

# Initialize traffic controller
if led_strip:
    traffic_controller = TrafficController(led_strip)
    traffic_controller.start()
    logger.info("Traffic controller started")
else:
    traffic_controller = None
    logger.warning("Traffic controller not initialized (LED strip unavailable)")

# Initialize DroidCam handler
droidcam = DroidCamHandler()
pedestrian_detector = PedestrianGestureDetector()


def gen_frames():
    """Generator function for streaming video frames with YOLO detection and traffic control"""
    global camera
    
    if not camera.is_running:
        if not camera.init_camera():
            logger.error("Failed to initialize camera for streaming")
            return
    
    # Load YOLO model on first request if enabled
    if camera.detector_enabled and not camera.detector.is_loaded:
        camera.detector.load_model()
    
    while True:
        frame = camera.get_frame()
        
        if frame is None:
            continue
        
        # Apply YOLO detection if enabled
        if camera.detector_enabled and camera.detector.is_loaded:
            # Use enhanced detection with ROI and tracking
            frame, direction_counts, tracked_objects = camera.detector.detect_vehicles(frame, draw_roi=True)
            
            # Update car count
            with camera.lock:
                camera.car_count = sum(direction_counts.values())
            
            # Get emergency vehicle info
            emergency_info = camera.detector.get_emergency_info()
            
            # Update traffic controller with vehicle counts and emergency info
            if traffic_controller:
                traffic_controller.update_vehicle_counts(direction_counts, emergency_info)
        else:
            # No detection, reset counts
            if traffic_controller:
                traffic_controller.update_vehicle_counts({d: 0 for d in TrafficController.DIRECTIONS})

        # Stream at up to 1080p (keeps aspect ratio, max width 1920)
        stream_frame = frame
        if frame.shape[1] > 1920:
            new_w = 1920
            new_h = int(frame.shape[0] * (1920.0 / frame.shape[1]))
            stream_frame = cv2.resize(frame, (new_w, new_h))

        # Encode frame to JPEG with higher quality for 1080p
        ret, buffer = cv2.imencode('.jpg', stream_frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
        
        if ret:
            frame_bytes = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n'
                   b'Content-Length: ' + str(len(frame_bytes)).encode() + b'\r\n\r\n' +
                   frame_bytes + b'\r\n')


def video_feed(request):
    """Stream video feed as MJPEG"""
    return StreamingHttpResponse(
        gen_frames(),
        content_type='multipart/x-mixed-replace; boundary=frame'
    )


def dashboard(request):
    """Display web dashboard with video feed"""
    return render(request, 'camera/dashboard.html', {'page': 'dashboard'})


def analytics(request):
    """Display analytics and logs page"""
    return render(request, 'camera/analytics.html', {'page': 'analytics'})


def cameras(request):
    """Display camera feeds page"""
    return render(request, 'camera/cameras.html', {'page': 'cameras'})


def settings_page(request):
    """Display settings page"""
    return render(request, 'camera/settings.html', {'page': 'settings'})


def camera_status(request):
    """API endpoint to check camera status"""
    global camera
    
    # Check if using picamera2 or OpenCV
    using_picamera2 = camera.picamera2 is not None
    using_opencv = camera.video is not None
    
    status = {
        'connected': camera.is_running,
        'camera_initialized': using_picamera2 or using_opencv,
        'backend': 'picamera2' if using_picamera2 else 'opencv' if using_opencv else 'none',
        'detection_enabled': camera.detector_enabled,
        'car_count': camera.car_count,
    }
    return JsonResponse(status)


@csrf_exempt
def toggle_detection(request):
    """Toggle YOLO car detection on/off"""
    global camera
    
    if request.method == 'POST':
        camera.detector_enabled = not camera.detector_enabled
        
        # Load model if enabling for first time
        if camera.detector_enabled and not camera.detector.is_loaded:
            success = camera.detector.load_model()
            return JsonResponse({
                'detection_enabled': camera.detector_enabled,
                'model_loaded': success,
                'message': 'Car detection enabled' if success else 'Failed to load model'
            })
        
        return JsonResponse({
            'detection_enabled': camera.detector_enabled,
            'message': 'Car detection enabled' if camera.detector_enabled else 'Car detection disabled'
        })
    
    return JsonResponse({'error': 'POST request required'}, status=400)


def detection_stats(request):
    """Get current detection statistics"""
    global camera
    
    stats = {
        'detection_enabled': camera.detector_enabled,
        'cars_detected': camera.car_count,
        'model_loaded': camera.detector.is_loaded,
    }
    return JsonResponse(stats)


@csrf_exempt
def set_traffic_mode(request):
    """Set traffic control mode (AUTO/MANUAL)"""
    if not traffic_controller:
        return JsonResponse({'error': 'Traffic controller not available'}, status=503)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            mode = data.get('mode', 'AUTO').upper()
            
            if mode not in ['AUTO', 'MANUAL']:
                return JsonResponse({'error': 'Invalid mode'}, status=400)
            
            success = traffic_controller.set_mode(mode)
            return JsonResponse({
                'success': success,
                'mode': mode,
                'message': f'Mode set to {mode}'
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'POST request required'}, status=400)


@csrf_exempt
def manual_control_light(request):
    """Manually control a traffic light (MANUAL mode only)"""
    if not traffic_controller:
        return JsonResponse({'error': 'Traffic controller not available'}, status=503)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            direction = data.get('direction')
            state = data.get('state', 'RED').upper()
            
            if not direction:
                return JsonResponse({'error': 'Direction required'}, status=400)
            
            success = traffic_controller.manual_set_direction(direction, state)
            
            if success:
                return JsonResponse({
                    'success': True,
                    'direction': direction,
                    'state': state
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Manual control requires MANUAL mode'
                }, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'POST request required'}, status=400)


@csrf_exempt
def request_pedestrian_crossing(request):
    """Request pedestrian crossing"""
    if not traffic_controller:
        return JsonResponse({'error': 'Traffic controller not available'}, status=503)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            direction = data.get('direction', 'NORTH')
            
            success = traffic_controller.request_pedestrian_crossing(direction)
            
            return JsonResponse({
                'success': success,
                'direction': direction,
                'message': 'Request accepted' if success else 'Request in cooldown'
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'POST request required'}, status=400)


def traffic_status(request):
    """Get complete traffic control system status"""
    if not traffic_controller:
        return JsonResponse({'error': 'Traffic controller not available'}, status=503)
    
    status = traffic_controller.get_status()
    
    # Add detector stats
    status['detector'] = {
        'enabled': camera.detector_enabled,
        'fps': camera.detector.get_fps(),
        'direction_counts': camera.detector.get_direction_counts()
    }
    
    return JsonResponse(status)


def traffic_detailed_status(request):
    """Get detailed traffic status with waiting times and priority scores"""
    if not traffic_controller:
        return JsonResponse({'error': 'Traffic controller not available'}, status=503)
    
    detailed = traffic_controller.get_detailed_status()
    return JsonResponse(detailed)


def get_algorithm_settings(request):
    """Get current algorithm settings"""
    if not traffic_controller:
        return JsonResponse({'error': 'Traffic controller not available'}, status=503)
    
    settings = traffic_controller.get_algorithm_settings()
    return JsonResponse({'settings': settings})


@csrf_exempt
def update_algorithm_settings(request):
    """Update algorithm settings"""
    if not traffic_controller:
        return JsonResponse({'error': 'Traffic controller not available'}, status=503)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required'}, status=400)
    
    try:
        data = json.loads(request.body)
        result = traffic_controller.update_algorithm_settings(data)
        return JsonResponse(result)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def event_log(request):
    """Get event log"""
    if not traffic_controller:
        return JsonResponse({'error': 'Traffic controller not available'}, status=503)
    
    limit = int(request.GET.get('limit', 50))
    events = traffic_controller.get_event_log(limit=limit)
    
    return JsonResponse({'events': events})


@csrf_exempt
def emergency_stop(request):
    """Emergency stop - set all lights to red"""
    if not traffic_controller:
        return JsonResponse({'error': 'Traffic controller not available'}, status=503)
    
    if request.method == 'POST':
        traffic_controller.emergency_stop()
        return JsonResponse({
            'success': True,
            'message': 'Emergency stop activated - all lights RED'
        })
    
    return JsonResponse({'error': 'POST request required'}, status=400)


@csrf_exempt
def test_led(request):
    """Test LED strip functionality"""
    if not led_strip:
        return JsonResponse({'error': 'LED strip not available'}, status=503)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            color = data.get('color', 'off').upper()
            brightness = data.get('brightness', None)
            
            # Set brightness if provided
            if brightness is not None:
                try:
                    led_strip.brightness = int(brightness) / 100.0
                except (ValueError, TypeError):
                    pass
            
            # Set the color/state
            valid_colors = ['RED', 'YELLOW', 'GREEN', 'ALL_ON', 'OFF']
            if color in valid_colors:
                led_strip.set_state(color)
                return JsonResponse({
                    'success': True,
                    'message': f'LED strip set to {color}',
                    'color': color,
                    'enabled': led_strip.enabled
                })
            else:
                return JsonResponse({
                    'error': f'Invalid color. Use: {valid_colors}'
                }, status=400)
                
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            logger.error(f"LED test error: {e}")
            return JsonResponse({'error': str(e)}, status=500)
    
    # GET request - return LED status
    return JsonResponse({
        'enabled': led_strip.enabled,
        'current_state': led_strip._current_state if hasattr(led_strip, '_current_state') else 'UNKNOWN',
        'num_pixels': led_strip.num_pixels,
        'brightness': int(led_strip.brightness * 100)
    })


@csrf_exempt
def start_droidcam(request):
    """Start DroidCam connection"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            url = data.get('url', 'http://192.168.1.100:4747/mjpegfeed')
            
            droidcam.droidcam_url = url
            success = droidcam.start()
            
            if success:
                # Load pedestrian detector model (shared with main YOLO)
                pedestrian_detector.load_model()
                
                return JsonResponse({
                    'success': True,
                    'message': 'DroidCam connected',
                    'url': url
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': droidcam.error_msg
                }, status=503)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'POST request required'}, status=400)


def droidcam_feed(request):
    """Stream DroidCam video feed with pedestrian gesture detection"""
    def gen_droidcam_frames():
        while droidcam.is_running:
            frame = droidcam.get_frame()
            
            if frame is None:
                continue
            
            # Apply pedestrian gesture detection
            if pedestrian_detector.is_loaded:
                gesture_detected, confidence, annotated_frame, direction = pedestrian_detector.detect_gesture(
                    frame, draw_overlay=True
                )
                
                if gesture_detected and direction and traffic_controller:
                    # Send crossing request
                    traffic_controller.request_pedestrian_crossing(direction)
                    logger.info(f"Pedestrian gesture detected - requesting crossing for {direction}")
                
                frame = annotated_frame
            
            # Encode frame
            ret, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 75])
            
            if ret:
                frame_bytes = buffer.tobytes()
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n'
                       b'Content-Length: ' + str(len(frame_bytes)).encode() + b'\r\n\r\n' +
                       frame_bytes + b'\r\n')
    
    return StreamingHttpResponse(
        gen_droidcam_frames(),
        content_type='multipart/x-mixed-replace; boundary=frame'
    )


def droidcam_status(request):
    """Get DroidCam status"""
    status = {
        'connected': droidcam.is_connected,
        'running': droidcam.is_running,
        'url': droidcam.droidcam_url,
        'error': droidcam.error_msg,
        'pedestrian_detector': pedestrian_detector.get_status()
    }
    
    return JsonResponse(status)


# Pedestrian Phone Mode state
pedestrian_phone_mode = {
    'enabled': False,
    'cars_detected': 0,
    'last_trigger': 0,
    'cooldown': 10  # seconds between triggers
}

def droidcam_pedestrian_mode(request):
    """
    Pedestrian Phone Mode - When DroidCam detects cars, trigger RED for cars (crossing for pedestrian)
    
    This mode works by:
    1. DroidCam captures video from pedestrian's phone
    2. YOLO detects cars in the frame
    3. If cars detected, system triggers RED for cars (pedestrian can cross)
    """
    import time
    from ultralytics import YOLO
    
    global pedestrian_phone_mode
    
    if not droidcam.is_running:
        return JsonResponse({
            'error': 'DroidCam not connected',
            'cars_detected': 0,
            'triggered': False
        })
    
    # Get current frame from DroidCam
    frame = droidcam.get_frame()
    
    if frame is None:
        return JsonResponse({
            'error': 'No frame available',
            'cars_detected': 0,
            'triggered': False
        })
    
    cars_detected = 0
    triggered = False
    
    try:
        # Use the main camera's detector to detect cars in DroidCam frame
        if camera and camera.detector:
            results = camera.detector.model(frame, verbose=False, conf=0.4)
            
            if results and len(results) > 0:
                result = results[0]
                
                # Count cars, trucks, buses, motorcycles (COCO classes: 2=car, 5=bus, 7=truck, 3=motorcycle)
                vehicle_classes = [2, 3, 5, 7]
                
                for box in result.boxes:
                    class_id = int(box.cls[0])
                    if class_id in vehicle_classes:
                        cars_detected += 1
        
        pedestrian_phone_mode['cars_detected'] = cars_detected
        
        # Check if we should trigger crossing
        current_time = time.time()
        time_since_last = current_time - pedestrian_phone_mode['last_trigger']
        
        if cars_detected > 0 and time_since_last > pedestrian_phone_mode['cooldown']:
            # Cars detected - trigger RED for cars (pedestrian crossing)
            if traffic_controller:
                # Request pedestrian crossing - this will make the light RED for cars
                traffic_controller.request_pedestrian_crossing('PEDESTRIAN_PHONE')
                pedestrian_phone_mode['last_trigger'] = current_time
                triggered = True
                logger.info(f"Pedestrian Phone Mode: {cars_detected} cars detected, triggering crossing")
                
                # Log to database
                try:
                    from detection.models import DetectionEvent
                    DetectionEvent.objects.create(
                        event_type='PEDESTRIAN',
                        message=f'Pedestrian Phone Mode: {cars_detected} cars detected - triggering crossing',
                        direction='PEDESTRIAN_PHONE',
                        value=cars_detected
                    )
                except Exception as e:
                    logger.error(f"Failed to log pedestrian phone event: {e}")
        
        return JsonResponse({
            'cars_detected': cars_detected,
            'triggered': triggered,
            'cooldown_remaining': max(0, pedestrian_phone_mode['cooldown'] - time_since_last),
            'mode_enabled': True
        })
        
    except Exception as e:
        logger.error(f"Pedestrian phone mode error: {e}")
        return JsonResponse({
            'error': str(e),
            'cars_detected': 0,
            'triggered': False
        })


@csrf_exempt
def shutdown_camera(request):
    """Shutdown camera gracefully"""
    global camera
    camera.release()
    
    if traffic_controller:
        traffic_controller.stop()
    
    if led_strip:
        led_strip.off()
    
    if droidcam.is_running:
        droidcam.stop()
    
    return JsonResponse({'status': 'System shutdown complete'})


@csrf_exempt
def backup_settings(request):
    """
    Backup current system settings to JSON
    Returns a downloadable JSON file
    """
    import os
    from datetime import datetime
    
    settings = {
        'version': '1.0',
        'backup_date': datetime.now().isoformat(),
        'traffic_controller': {},
        'detection': {},
        'led_strip': {},
        'zones': {}
    }
    
    # Traffic controller settings
    if traffic_controller:
        settings['traffic_controller'] = {
            'mode': traffic_controller.mode,
            'T_MIN': traffic_controller.T_MIN,
            'T_MAX': traffic_controller.T_MAX,
            'T_PER_VEHICLE': traffic_controller.T_PER_VEHICLE,
            'T_YELLOW': traffic_controller.T_YELLOW,
            'T_PEDESTRIAN': traffic_controller.T_PEDESTRIAN,
            'SIMPLE_GREEN_DURATION': traffic_controller.SIMPLE_GREEN_DURATION,
            'SIMPLE_YELLOW_DURATION': traffic_controller.SIMPLE_YELLOW_DURATION,
            'EMERGENCY_PRIORITY': traffic_controller.EMERGENCY_PRIORITY,
            'EMERGENCY_GREEN_TIME': traffic_controller.EMERGENCY_GREEN_TIME,
        }
    
    # Detection settings
    if camera and camera.detector:
        settings['detection'] = {
            'enabled': camera.detector_enabled,
            'confidence_threshold': camera.detector.confidence_threshold,
        }
    
    # LED strip settings
    if led_strip:
        settings['led_strip'] = {
            'num_pixels': led_strip.num_pixels,
            'brightness': led_strip.brightness,
            'current_state': led_strip.get_state() if hasattr(led_strip, 'get_state') else 'UNKNOWN',
        }
    
    # Zone configurations
    if camera and camera.detector:
        settings['zones'] = camera.detector.roi_zones
    
    return JsonResponse(settings)


@csrf_exempt
def restore_settings(request):
    """
    Restore system settings from JSON backup
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST request required'}, status=400)
    
    try:
        data = json.loads(request.body)
        
        restored = []
        
        # Restore traffic controller settings
        if traffic_controller and 'traffic_controller' in data:
            tc = data['traffic_controller']
            if 'mode' in tc:
                traffic_controller.set_mode(tc['mode'])
                restored.append('mode')
            if 'T_MIN' in tc:
                traffic_controller.T_MIN = tc['T_MIN']
                restored.append('T_MIN')
            if 'T_MAX' in tc:
                traffic_controller.T_MAX = tc['T_MAX']
                restored.append('T_MAX')
            if 'T_PER_VEHICLE' in tc:
                traffic_controller.T_PER_VEHICLE = tc['T_PER_VEHICLE']
                restored.append('T_PER_VEHICLE')
            if 'SIMPLE_GREEN_DURATION' in tc:
                traffic_controller.SIMPLE_GREEN_DURATION = tc['SIMPLE_GREEN_DURATION']
                restored.append('SIMPLE_GREEN_DURATION')
            if 'SIMPLE_YELLOW_DURATION' in tc:
                traffic_controller.SIMPLE_YELLOW_DURATION = tc['SIMPLE_YELLOW_DURATION']
                restored.append('SIMPLE_YELLOW_DURATION')
            if 'EMERGENCY_PRIORITY' in tc:
                traffic_controller.EMERGENCY_PRIORITY = tc['EMERGENCY_PRIORITY']
                restored.append('EMERGENCY_PRIORITY')
            if 'EMERGENCY_GREEN_TIME' in tc:
                traffic_controller.EMERGENCY_GREEN_TIME = tc['EMERGENCY_GREEN_TIME']
                restored.append('EMERGENCY_GREEN_TIME')
        
        # Restore detection settings
        if camera and camera.detector and 'detection' in data:
            det = data['detection']
            if 'enabled' in det:
                camera.detector_enabled = det['enabled']
                restored.append('detection_enabled')
            if 'confidence_threshold' in det:
                camera.detector.confidence_threshold = det['confidence_threshold']
                restored.append('confidence_threshold')
        
        # Restore LED settings
        if led_strip and 'led_strip' in data:
            ls = data['led_strip']
            if 'brightness' in ls:
                led_strip.brightness = ls['brightness']
                restored.append('brightness')
        
        # Restore zone configurations
        if camera and camera.detector and 'zones' in data:
            for direction, coords in data['zones'].items():
                if direction in camera.detector.roi_zones:
                    camera.detector.roi_zones[direction] = tuple(coords) if isinstance(coords, list) else coords
                    restored.append(f'zone_{direction}')
        
        logger.info(f"Settings restored: {restored}")
        
        return JsonResponse({
            'success': True,
            'message': f'Restored {len(restored)} settings',
            'restored': restored
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"Restore settings error: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def configure_zones(request):
    """
    Configure detection zones for single camera multi-direction counting
    """
    if request.method == 'GET':
        # Return current zone configuration
        if camera and camera.detector:
            return JsonResponse({
                'zones': camera.detector.roi_zones,
                'supported_directions': ['NORTH', 'EAST', 'SOUTH', 'WEST']
            })
        return JsonResponse({'error': 'Detector not available'}, status=503)
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            if not camera or not camera.detector:
                return JsonResponse({'error': 'Detector not available'}, status=503)
            
            # Update zones
            for direction, zone in data.items():
                if direction in camera.detector.roi_zones:
                    if isinstance(zone, dict):
                        camera.detector.roi_zones[direction] = (
                            zone.get('x1', 0),
                            zone.get('y1', 0),
                            zone.get('x2', 1),
                            zone.get('y2', 1)
                        )
                    elif isinstance(zone, (list, tuple)) and len(zone) == 4:
                        camera.detector.roi_zones[direction] = tuple(zone)
            
            logger.info(f"Zones updated: {camera.detector.roi_zones}")
            
            return JsonResponse({
                'success': True,
                'zones': camera.detector.roi_zones
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ==================== ANALYTICS API ENDPOINTS ====================

def vehicle_count_history(request):
    """
    Get vehicle count history from database for graphs
    Query params:
        - hours: Number of hours to look back (default: 24)
        - limit: Max records to return (default: 500)
    """
    from detection.models import VehicleCount
    from django.utils import timezone
    from datetime import timedelta
    
    hours = int(request.GET.get('hours', 24))
    limit = int(request.GET.get('limit', 500))
    
    since = timezone.now() - timedelta(hours=hours)
    
    records = VehicleCount.objects.filter(
        timestamp__gte=since
    ).order_by('timestamp')[:limit]
    
    data = {
        'labels': [],
        'north': [],
        'east': [],
        'south': [],
        'west': [],
        'total': []
    }
    
    for record in records:
        data['labels'].append(record.timestamp.strftime('%H:%M:%S'))
        data['north'].append(record.north_count)
        data['east'].append(record.east_count)
        data['south'].append(record.south_count)
        data['west'].append(record.west_count)
        data['total'].append(record.total_count)
    
    return JsonResponse(data)


def daily_stats(request):
    """
    Get daily statistics for the past N days
    Query params:
        - days: Number of days to look back (default: 7)
    """
    from detection.models import SystemStats
    from datetime import date, timedelta
    
    days = int(request.GET.get('days', 7))
    
    since = date.today() - timedelta(days=days)
    
    records = SystemStats.objects.filter(
        date__gte=since
    ).order_by('date')
    
    data = {
        'labels': [],
        'vehicles': [],
        'pedestrians': [],
        'cycles': []
    }
    
    for record in records:
        data['labels'].append(record.date.strftime('%Y-%m-%d'))
        data['vehicles'].append(record.total_vehicles_detected)
        data['pedestrians'].append(record.total_pedestrian_requests)
        data['cycles'].append(record.total_light_cycles)
    
    return JsonResponse(data)


def led_change_history(request):
    """
    Get LED state change history
    Query params:
        - hours: Number of hours to look back (default: 24)
        - limit: Max records to return (default: 200)
    """
    from detection.models import TrafficLightState
    from django.utils import timezone
    from datetime import timedelta
    
    hours = int(request.GET.get('hours', 24))
    limit = int(request.GET.get('limit', 200))
    
    since = timezone.now() - timedelta(hours=hours)
    
    records = TrafficLightState.objects.filter(
        timestamp__gte=since
    ).order_by('-timestamp')[:limit]
    
    data = {
        'changes': [],
        'state_counts': {'RED': 0, 'YELLOW': 0, 'GREEN': 0, 'RED_YELLOW': 0}
    }
    
    for record in records:
        data['changes'].append({
            'timestamp': record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'state': record.state,
            'direction': record.direction,
            'triggered_by': record.triggered_by
        })
        if record.state in data['state_counts']:
            data['state_counts'][record.state] += 1
    
    return JsonResponse(data)


def analytics_summary(request):
    """
    Get a summary of all analytics data for dashboard
    """
    from detection.models import DetectionEvent, VehicleCount, TrafficLightState, SystemStats
    from django.utils import timezone
    from django.db.models import Sum, Count, Avg
    from datetime import timedelta, date
    
    # Time ranges
    now = timezone.now()
    today = date.today()
    last_hour = now - timedelta(hours=1)
    last_24h = now - timedelta(hours=24)
    
    # Today's stats
    today_stats = SystemStats.objects.filter(date=today).first()
    
    # Last hour vehicle counts
    last_hour_counts = VehicleCount.objects.filter(
        timestamp__gte=last_hour
    ).aggregate(
        total=Sum('total_count'),
        avg_north=Avg('north_count'),
        avg_east=Avg('east_count'),
        avg_south=Avg('south_count'),
        avg_west=Avg('west_count')
    )
    
    # LED state distribution (last 24h)
    led_states = TrafficLightState.objects.filter(
        timestamp__gte=last_24h
    ).values('state').annotate(count=Count('id'))
    
    led_distribution = {item['state']: item['count'] for item in led_states}
    
    # Event counts by type (last 24h)
    event_counts = DetectionEvent.objects.filter(
        timestamp__gte=last_24h
    ).values('event_type').annotate(count=Count('id'))
    
    events_by_type = {item['event_type']: item['count'] for item in event_counts}
    
    # Peak hour analysis (which hour had most vehicles)
    peak_hour_data = VehicleCount.objects.filter(
        timestamp__gte=last_24h
    ).extra(
        select={'hour': 'strftime("%%H", timestamp)'}
    ).values('hour').annotate(
        total=Sum('total_count')
    ).order_by('-total').first()
    
    summary = {
        'today': {
            'total_vehicles': today_stats.total_vehicles_detected if today_stats else 0,
            'pedestrian_requests': today_stats.total_pedestrian_requests if today_stats else 0,
            'light_cycles': today_stats.total_light_cycles if today_stats else 0
        },
        'last_hour': {
            'total_vehicles': last_hour_counts['total'] or 0,
            'avg_by_direction': {
                'NORTH': round(last_hour_counts['avg_north'] or 0, 1),
                'EAST': round(last_hour_counts['avg_east'] or 0, 1),
                'SOUTH': round(last_hour_counts['avg_south'] or 0, 1),
                'WEST': round(last_hour_counts['avg_west'] or 0, 1)
            }
        },
        'led_distribution': led_distribution,
        'events_by_type': events_by_type,
        'peak_hour': peak_hour_data['hour'] if peak_hour_data else None,
        'database_records': {
            'detection_events': DetectionEvent.objects.count(),
            'vehicle_counts': VehicleCount.objects.count(),
            'led_states': TrafficLightState.objects.count(),
            'daily_stats': SystemStats.objects.count()
        }
    }
    
    return JsonResponse(summary)


# ==================== PEAK HOUR ANALYSIS ====================

def peak_hour_analysis(request):
    """
    Analyze traffic patterns by hour to find peak hours
    Query params:
        - days: Number of days to analyze (default: 7)
    """
    from detection.models import VehicleCount, HourlyStats
    from django.utils import timezone
    from django.db.models import Sum, Avg
    from datetime import timedelta, date
    
    days = int(request.GET.get('days', 7))
    since = timezone.now() - timedelta(days=days)
    
    # Aggregate by hour
    hourly_data = VehicleCount.objects.filter(
        timestamp__gte=since
    ).extra(
        select={'hour': 'strftime("%%H", timestamp)'}
    ).values('hour').annotate(
        total=Sum('total_count'),
        avg_total=Avg('total_count'),
        north=Sum('north_count'),
        east=Sum('east_count'),
        south=Sum('south_count'),
        west=Sum('west_count')
    ).order_by('hour')
    
    # Create 24-hour array
    hours = {str(i).zfill(2): {'total': 0, 'avg': 0, 'north': 0, 'east': 0, 'south': 0, 'west': 0} for i in range(24)}
    
    for item in hourly_data:
        hour = item['hour']
        if hour:
            hours[hour] = {
                'total': item['total'] or 0,
                'avg': round(item['avg_total'] or 0, 1),
                'north': item['north'] or 0,
                'east': item['east'] or 0,
                'south': item['south'] or 0,
                'west': item['west'] or 0
            }
    
    # Find peak hours
    sorted_hours = sorted(hours.items(), key=lambda x: x[1]['total'], reverse=True)
    peak_hours = [h[0] for h in sorted_hours[:3]]  # Top 3 peak hours
    quiet_hours = [h[0] for h in sorted_hours[-3:]]  # 3 quietest hours
    
    return JsonResponse({
        'hourly_data': hours,
        'peak_hours': peak_hours,
        'quiet_hours': quiet_hours,
        'analysis_period_days': days,
        'labels': [f"{i:02d}:00" for i in range(24)],
        'totals': [hours[f"{i:02d}"]['total'] for i in range(24)]
    })


# ==================== HEATMAP DATA ====================

def traffic_heatmap(request):
    """
    Generate heatmap data for traffic visualization
    Returns a matrix of hour x day-of-week with traffic intensity
    Query params:
        - days: Number of days to analyze (default: 28)
    """
    from detection.models import VehicleCount
    from django.utils import timezone
    from django.db.models import Sum
    from datetime import timedelta
    
    days = int(request.GET.get('days', 28))
    since = timezone.now() - timedelta(days=days)
    
    # Get data grouped by day of week and hour
    data = VehicleCount.objects.filter(
        timestamp__gte=since
    ).extra(
        select={
            'hour': 'strftime("%%H", timestamp)',
            'weekday': 'strftime("%%w", timestamp)'  # 0=Sunday, 6=Saturday
        }
    ).values('hour', 'weekday').annotate(
        total=Sum('total_count')
    )
    
    # Create 7x24 matrix (days x hours)
    day_names = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    heatmap = [[0 for _ in range(24)] for _ in range(7)]
    max_value = 0
    
    for item in data:
        if item['hour'] and item['weekday']:
            hour = int(item['hour'])
            weekday = int(item['weekday'])
            value = item['total'] or 0
            heatmap[weekday][hour] = value
            max_value = max(max_value, value)
    
    # Normalize to 0-100 for color intensity
    normalized = [[0 for _ in range(24)] for _ in range(7)]
    if max_value > 0:
        for d in range(7):
            for h in range(24):
                normalized[d][h] = round((heatmap[d][h] / max_value) * 100, 1)
    
    return JsonResponse({
        'raw_data': heatmap,
        'normalized': normalized,
        'max_value': max_value,
        'day_labels': day_names,
        'hour_labels': [f"{i:02d}:00" for i in range(24)],
        'analysis_period_days': days
    })


# ==================== EXPORT TO CSV/EXCEL ====================

def export_csv(request):
    """
    Export traffic data to CSV
    Query params:
        - type: 'vehicles', 'events', 'led_states' (default: vehicles)
        - days: Number of days to export (default: 7)
    """
    from detection.models import VehicleCount, DetectionEvent, TrafficLightState
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import timedelta
    import csv
    
    export_type = request.GET.get('type', 'vehicles')
    days = int(request.GET.get('days', 7))
    since = timezone.now() - timedelta(days=days)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="traffic_{export_type}_{days}days.csv"'
    
    writer = csv.writer(response)
    
    if export_type == 'vehicles':
        writer.writerow(['Timestamp', 'North', 'East', 'South', 'West', 'Total'])
        records = VehicleCount.objects.filter(timestamp__gte=since).order_by('timestamp')
        for r in records:
            writer.writerow([r.timestamp.isoformat(), r.north_count, r.east_count, r.south_count, r.west_count, r.total_count])
    
    elif export_type == 'events':
        writer.writerow(['Timestamp', 'Event Type', 'Direction', 'Vehicle Count', 'Message'])
        records = DetectionEvent.objects.filter(timestamp__gte=since).order_by('timestamp')
        for r in records:
            writer.writerow([r.timestamp.isoformat(), r.event_type, r.direction, r.vehicle_count, r.message])
    
    elif export_type == 'led_states':
        writer.writerow(['Timestamp', 'State', 'Direction', 'Triggered By'])
        records = TrafficLightState.objects.filter(timestamp__gte=since).order_by('timestamp')
        for r in records:
            writer.writerow([r.timestamp.isoformat(), r.state, r.direction, r.triggered_by])
    
    return response


def export_excel(request):
    """
    Export traffic data to Excel format
    Query params:
        - days: Number of days to export (default: 7)
    """
    from detection.models import VehicleCount, DetectionEvent, TrafficLightState, SystemStats
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import timedelta
    
    days = int(request.GET.get('days', 7))
    since = timezone.now() - timedelta(days=days)
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        
        # Vehicle Counts Sheet
        ws1 = wb.active
        ws1.title = "Vehicle Counts"
        headers = ['Timestamp', 'North', 'East', 'South', 'West', 'Total']
        for col, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        records = VehicleCount.objects.filter(timestamp__gte=since).order_by('timestamp')
        for row, r in enumerate(records, 2):
            ws1.cell(row=row, column=1, value=r.timestamp.isoformat())
            ws1.cell(row=row, column=2, value=r.north_count)
            ws1.cell(row=row, column=3, value=r.east_count)
            ws1.cell(row=row, column=4, value=r.south_count)
            ws1.cell(row=row, column=5, value=r.west_count)
            ws1.cell(row=row, column=6, value=r.total_count)
        
        # Events Sheet
        ws2 = wb.create_sheet("Detection Events")
        headers = ['Timestamp', 'Event Type', 'Direction', 'Vehicle Count', 'Message']
        for col, header in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        events = DetectionEvent.objects.filter(timestamp__gte=since).order_by('timestamp')
        for row, e in enumerate(events, 2):
            ws2.cell(row=row, column=1, value=e.timestamp.isoformat())
            ws2.cell(row=row, column=2, value=e.event_type)
            ws2.cell(row=row, column=3, value=e.direction)
            ws2.cell(row=row, column=4, value=e.vehicle_count)
            ws2.cell(row=row, column=5, value=e.message)
        
        # Daily Stats Sheet
        ws3 = wb.create_sheet("Daily Stats")
        headers = ['Date', 'Vehicles', 'Pedestrians', 'Light Cycles', 'Emergencies']
        for col, header in enumerate(headers, 1):
            ws3.cell(row=1, column=col, value=header).font = Font(bold=True)
        
        stats = SystemStats.objects.all().order_by('-date')[:days]
        for row, s in enumerate(stats, 2):
            ws3.cell(row=row, column=1, value=str(s.date))
            ws3.cell(row=row, column=2, value=s.total_vehicles_detected)
            ws3.cell(row=row, column=3, value=s.total_pedestrian_requests)
            ws3.cell(row=row, column=4, value=s.total_light_cycles)
            ws3.cell(row=row, column=5, value=s.total_emergency_stops)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="traffic_report_{days}days.xlsx"'
        wb.save(response)
        return response
        
    except ImportError:
        return JsonResponse({
            'error': 'openpyxl not installed. Install with: pip install openpyxl',
            'fallback': 'Use CSV export instead'
        }, status=500)


# ==================== WEATHER INTEGRATION ====================

def get_weather(request):
    """
    Get current weather data from OpenWeatherMap API
    Configure WEATHER_API_KEY and WEATHER_CITY in settings
    Query params:
        - city: City name (default from settings or 'Bucharest')
        - lat: Latitude (optional, overrides city)
        - lon: Longitude (optional, overrides city)
    """
    from detection.models import WeatherData
    from django.conf import settings
    import os
    
    # Get API key from settings or environment
    api_key = getattr(settings, 'WEATHER_API_KEY', os.environ.get('WEATHER_API_KEY', ''))
    
    if not api_key:
        return JsonResponse({
            'error': 'Weather API key not configured',
            'help': 'Set WEATHER_API_KEY in settings.py or as environment variable',
            'demo_mode': True,
            'weather': {
                'temperature': 15,
                'humidity': 65,
                'condition': 'Clear',
                'description': 'Demo mode - API key not set',
                'wind_speed': 5
            }
        })
    
    # Get city from settings or request
    default_city = getattr(settings, 'WEATHER_CITY', 'Bucharest')
    city = request.GET.get('city', default_city)
    lat = request.GET.get('lat')
    lon = request.GET.get('lon')
    
    try:
        import requests
        
        # Use lat/lon if provided, otherwise use city name
        if lat and lon:
            url = f"https://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&appid={api_key}&units=metric"
        else:
            url = f"https://api.openweathermap.org/data/2.5/weather?q={city}&appid={api_key}&units=metric"
        
        response = requests.get(url, timeout=10)
        data = response.json()
        
        if response.status_code == 200:
            weather_info = {
                'temperature': data['main']['temp'],
                'humidity': data['main']['humidity'],
                'condition': data['weather'][0]['main'],
                'description': data['weather'][0]['description'],
                'wind_speed': data['wind']['speed'],
                'visibility': data.get('visibility', 10000)
            }
            
            # Store in database
            WeatherData.objects.create(
                temperature=weather_info['temperature'],
                humidity=weather_info['humidity'],
                weather_condition=weather_info['condition'],
                weather_description=weather_info['description'],
                wind_speed=weather_info['wind_speed'],
                visibility=weather_info['visibility']
            )
            
            return JsonResponse({'weather': weather_info, 'source': 'OpenWeatherMap'})
        else:
            return JsonResponse({'error': data.get('message', 'API error')}, status=400)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def weather_traffic_correlation(request):
    """
    Analyze correlation between weather and traffic
    Query params:
        - days: Number of days to analyze (default: 7)
    """
    from detection.models import VehicleCount, WeatherData
    from django.utils import timezone
    from django.db.models import Avg, Sum
    from datetime import timedelta
    
    days = int(request.GET.get('days', 7))
    since = timezone.now() - timedelta(days=days)
    
    # Get weather data grouped by condition
    weather_groups = WeatherData.objects.filter(
        timestamp__gte=since
    ).values('weather_condition').annotate(
        avg_temp=Avg('temperature'),
        count=Sum('id')
    )
    
    # Get traffic during each weather condition
    correlation = {}
    for weather in weather_groups:
        condition = weather['weather_condition']
        
        # Get timestamps for this weather condition
        weather_times = WeatherData.objects.filter(
            timestamp__gte=since,
            weather_condition=condition
        ).values_list('timestamp', flat=True)
        
        # Approximate: get vehicle counts near weather readings
        if weather_times:
            avg_vehicles = VehicleCount.objects.filter(
                timestamp__gte=since
            ).aggregate(avg=Avg('total_count'))
            
            correlation[condition] = {
                'avg_temperature': round(weather['avg_temp'] or 0, 1),
                'occurrences': len(weather_times),
                'avg_vehicles': round(avg_vehicles['avg'] or 0, 1)
            }
    
    return JsonResponse({
        'correlation': correlation,
        'analysis_period_days': days,
        'insight': 'More data needed for accurate correlation' if len(correlation) < 3 else 'Correlation data available'
    })


# ==================== MULTI-CAMERA SUPPORT ====================

@csrf_exempt
def list_cameras(request):
    """List all registered camera sources"""
    from detection.models import CameraSource
    
    cameras = CameraSource.objects.all()
    
    return JsonResponse({
        'cameras': [
            {
                'id': cam.id,
                'name': cam.name,
                'type': cam.camera_type,
                'location': cam.location,
                'url': cam.url,
                'is_active': cam.is_active,
                'primary_direction': cam.primary_direction
            }
            for cam in cameras
        ],
        'total': cameras.count()
    })


@csrf_exempt
def add_camera(request):
    """Add a new camera source"""
    from detection.models import CameraSource
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    try:
        data = json.loads(request.body)
        
        camera = CameraSource.objects.create(
            name=data.get('name', 'New Camera'),
            camera_type=data.get('type', 'IP'),
            location=data.get('location', ''),
            url=data.get('url', ''),
            is_active=data.get('is_active', True),
            primary_direction=data.get('primary_direction', 'NORTH')
        )
        
        return JsonResponse({
            'success': True,
            'camera_id': camera.id,
            'message': f'Camera "{camera.name}" added successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def aggregate_camera_data(request):
    """
    Aggregate vehicle counts from all cameras
    Query params:
        - hours: Number of hours to look back (default: 24)
    """
    from detection.models import CameraSource, CameraVehicleCount, VehicleCount
    from django.utils import timezone
    from django.db.models import Sum
    from datetime import timedelta
    
    hours = int(request.GET.get('hours', 24))
    since = timezone.now() - timedelta(hours=hours)
    
    # Get per-camera aggregates
    cameras = CameraSource.objects.filter(is_active=True)
    camera_data = []
    
    for cam in cameras:
        counts = CameraVehicleCount.objects.filter(
            camera=cam,
            timestamp__gte=since
        ).aggregate(total=Sum('vehicle_count'))
        
        camera_data.append({
            'camera_id': cam.id,
            'name': cam.name,
            'location': cam.location,
            'direction': cam.primary_direction,
            'total_vehicles': counts['total'] or 0
        })
    
    # Get main camera (RPi) data
    main_data = VehicleCount.objects.filter(
        timestamp__gte=since
    ).aggregate(
        north=Sum('north_count'),
        east=Sum('east_count'),
        south=Sum('south_count'),
        west=Sum('west_count'),
        total=Sum('total_count')
    )
    
    return JsonResponse({
        'main_camera': {
            'north': main_data['north'] or 0,
            'east': main_data['east'] or 0,
            'south': main_data['south'] or 0,
            'west': main_data['west'] or 0,
            'total': main_data['total'] or 0
        },
        'additional_cameras': camera_data,
        'hours_analyzed': hours
    })


# ==================== TRAFFIC PREDICTION (ML-BASED) ====================

def predict_next_hour(request):
    """
    Predict traffic for the next hour using pattern-based ML
    Uses historical data patterns to make predictions
    """
    from detection.models import VehicleCount, HourlyStats, TrafficPrediction
    from django.utils import timezone
    from django.db.models import Avg
    from datetime import timedelta, datetime
    import statistics
    
    now = timezone.now()
    next_hour = now.hour + 1 if now.hour < 23 else 0
    
    # Get historical data for the same hour over past weeks
    same_hour_data = []
    for weeks_back in range(1, 5):  # Last 4 weeks
        target_time = now - timedelta(weeks=weeks_back)
        hour_start = target_time.replace(hour=next_hour, minute=0, second=0, microsecond=0)
        hour_end = hour_start + timedelta(hours=1)
        
        data = VehicleCount.objects.filter(
            timestamp__gte=hour_start,
            timestamp__lt=hour_end
        ).aggregate(
            north=Avg('north_count'),
            east=Avg('east_count'),
            south=Avg('south_count'),
            west=Avg('west_count')
        )
        
        if data['north'] is not None:
            same_hour_data.append(data)
    
    # Also get recent trend (last 3 hours)
    recent_data = VehicleCount.objects.filter(
        timestamp__gte=now - timedelta(hours=3)
    ).aggregate(
        north=Avg('north_count'),
        east=Avg('east_count'),
        south=Avg('south_count'),
        west=Avg('west_count')
    )
    
    # Combine historical patterns with recent trends (70% historical, 30% recent)
    if same_hour_data:
        hist_north = statistics.mean([d['north'] or 0 for d in same_hour_data])
        hist_east = statistics.mean([d['east'] or 0 for d in same_hour_data])
        hist_south = statistics.mean([d['south'] or 0 for d in same_hour_data])
        hist_west = statistics.mean([d['west'] or 0 for d in same_hour_data])
        
        pred_north = int(0.7 * hist_north + 0.3 * (recent_data['north'] or 0))
        pred_east = int(0.7 * hist_east + 0.3 * (recent_data['east'] or 0))
        pred_south = int(0.7 * hist_south + 0.3 * (recent_data['south'] or 0))
        pred_west = int(0.7 * hist_west + 0.3 * (recent_data['west'] or 0))
        confidence = min(0.9, 0.5 + len(same_hour_data) * 0.1)
    else:
        # Fallback to recent data only
        pred_north = int(recent_data['north'] or 0)
        pred_east = int(recent_data['east'] or 0)
        pred_south = int(recent_data['south'] or 0)
        pred_west = int(recent_data['west'] or 0)
        confidence = 0.3  # Low confidence without historical data
    
    total = pred_north + pred_east + pred_south + pred_west
    
    # Store prediction
    prediction_time = now.replace(hour=next_hour, minute=0, second=0, microsecond=0)
    if now.hour == 23:
        prediction_time += timedelta(days=1)
    
    try:
        TrafficPrediction.objects.create(
            prediction_for=prediction_time,
            predicted_north=pred_north,
            predicted_east=pred_east,
            predicted_south=pred_south,
            predicted_west=pred_west,
            predicted_total=total,
            confidence=confidence,
            model_version='pattern_v1'
        )
    except Exception as e:
        logger.debug(f"Could not save prediction: {e}")
    
    return JsonResponse({
        'prediction_for': f"{next_hour:02d}:00",
        'predictions': {
            'north': pred_north,
            'east': pred_east,
            'south': pred_south,
            'west': pred_west,
            'total': total
        },
        'confidence': round(confidence, 2),
        'model': 'pattern_v1',
        'data_points_used': len(same_hour_data),
        'note': 'Based on historical patterns for this hour of day'
    })


def predict_daily(request):
    """
    Predict traffic for the next 24 hours
    """
    from detection.models import VehicleCount
    from django.utils import timezone
    from django.db.models import Avg
    from datetime import timedelta
    import statistics
    
    now = timezone.now()
    predictions = []
    
    for hour_offset in range(1, 25):
        target_hour = (now.hour + hour_offset) % 24
        
        # Get historical data for this hour (last 4 weeks)
        same_hour_data = []
        for weeks_back in range(1, 5):
            target_time = now - timedelta(weeks=weeks_back)
            hour_start = target_time.replace(hour=target_hour, minute=0, second=0, microsecond=0)
            hour_end = hour_start + timedelta(hours=1)
            
            data = VehicleCount.objects.filter(
                timestamp__gte=hour_start,
                timestamp__lt=hour_end
            ).aggregate(total=Avg('total_count'))
            
            if data['total'] is not None:
                same_hour_data.append(data['total'])
        
        if same_hour_data:
            predicted = int(statistics.mean(same_hour_data))
            confidence = min(0.9, 0.5 + len(same_hour_data) * 0.1)
        else:
            # Fallback: estimate based on typical patterns
            if 7 <= target_hour <= 9 or 17 <= target_hour <= 19:
                predicted = 10  # Rush hour estimate
            elif 22 <= target_hour or target_hour <= 5:
                predicted = 1  # Night time estimate
            else:
                predicted = 5  # Normal hours estimate
            confidence = 0.2
        
        predictions.append({
            'hour': f"{target_hour:02d}:00",
            'predicted_total': predicted,
            'confidence': round(confidence, 2)
        })
    
    # Find predicted peak hours
    sorted_pred = sorted(predictions, key=lambda x: x['predicted_total'], reverse=True)
    
    return JsonResponse({
        'predictions': predictions,
        'peak_hours': [p['hour'] for p in sorted_pred[:3]],
        'quiet_hours': [p['hour'] for p in sorted_pred[-3:]],
        'model': 'daily_pattern_v1',
        'generated_at': now.isoformat()
    })